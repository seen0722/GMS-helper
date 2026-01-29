from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc
from backend.database.database import get_db
from backend.database import models
import pandas as pd
from io import BytesIO
from datetime import datetime

router = APIRouter()

@router.get("/submission/{submission_id}/excel")
def export_submission_excel(submission_id: int, db: Session = Depends(get_db)):
    """Export enhanced submission details to Excel (v2)."""
    from backend.services.merge_service import MergeService
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, colors
    from openpyxl.utils import get_column_letter
    from dateutil import parser as date_parser

    # Helper to format time
    def clean_time(time_str):
        if not time_str: return ""
        try:
            dt = date_parser.parse(time_str, fuzzy=True)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return time_str

    # 1. Fetch Consolidated Data
    report_data = MergeService.get_merge_report(db, submission_id)
    if not report_data:
        raise HTTPException(status_code=404, detail="Submission not found or no runs")
        
    sub = report_data['submission']
    
    # 2. Prepare Sheets Data
    
    # --- Calculate Overall Effective Pass Rate ---
    total_tests_all = 0
    total_remaining_all = report_data['remaining_failures']
    
    # Suite Stats Table
    suite_headers = ["Suite", "Initial Fail", "Recovered (Flaky)", "Remaining (Actionable)", "Pass Rate"]
    suite_rows = []
    
    for suite in report_data['suites']:
        s_sum = suite['summary']
        latest_run = suite['runs'][-1]
        
        # Determine total tests for this suite logic
        # PRE-FIXED in MergeService to be consistent with UI (Max of Passed + Failed)
        run_total = suite.get('total_tests', 0)
        total_tests_all += run_total
        
        eff_pass_rate = 100.0
        if run_total > 0:
            eff_pass_rate = max(0, (run_total - s_sum['remaining'])) / run_total * 100
            
        suite_rows.append([
            suite['suite_name'],
            s_sum['initial'],
            s_sum['recovered'],
            s_sum['remaining'],
            f"{eff_pass_rate:.4f}%"
        ])
        
    # Overall Calculation
    overall_pass_rate_str = "N/A"
    if total_tests_all > 0:
        overall_val = ((total_tests_all - total_remaining_all) / total_tests_all) * 100
        overall_pass_rate_str = f"{overall_val:.4f}%"

    # --- Sheet 1: Executive Summary ---
    summary_rows = [
        ["Submission", sub.name],
        ["Target Build", sub.target_fingerprint],
        ["GMS Version", sub.gms_version],
        ["Status", sub.status],
        ["Pass Rate (Effective)", overall_pass_rate_str],
        ["", ""],
        ["Overall Health", ""],
        ["Active Failures (Persistent)", report_data['remaining_failures']],
        ["Recovered Failures (Flaky)", report_data['total_recovered']],
        ["Total Initial Failures", report_data['total_initial_failures']],
        ["", ""],
        ["Suite Breakdown", ""]  # Header for next table
    ]

    df_summary_kpi = pd.DataFrame(summary_rows)
    df_summary_suites = pd.DataFrame(suite_rows, columns=suite_headers)

    # --- Sheet 2: Consolidated Analysis (The Worksheet) ---
    cons_rows = []
    base_url = "http://localhost:8000" # TODO: Get from settings
    
    for suite in report_data['suites']:
        for item in suite['items']:
            # Determine Owner
            owner = "Android Team" # Default
            if "Camera" in item['module_name']: owner = "Media/Camera Team"
            elif "Audio" in item['module_name']: owner = "Media/Audio Team"
            elif "Net" in item['module_name'] or "Radio" in item['module_name']: owner = "Connectivity Team"
            elif "Security" in item['module_name']: owner = "Security Team"
            
            # Formatting Status
            final_status = "RECOVERED" if item['is_recovered'] else "PERSISTENT"
            
            # Formatting History
            hist_str = " -> ".join([s.upper() for s in item['status_history']])
            
            # Stability Score
            fail_count = item['status_history'].count('fail')
            total_runs = len(item['status_history'])
            stability = f"Unstable ({fail_count}/{total_runs} failed)" if item['is_recovered'] else "Consistent Fail"
            
            # AI & Web Link
            f_obj = item.get('failure_details')
            ai_cause = ""
            bug_id = ""
            if f_obj and f_obj.failure_analysis:
                ai_cause = f_obj.failure_analysis.root_cause or ""
            
            web_link = f"{base_url}/submissions/{submission_id}/runs/{item['final_run_id']}?case={f_obj.id}" if f_obj else ""
            
            cons_rows.append({
                "Suite": suite['suite_name'],
                "Module": item['module_name'],
                "ABI": item.get('module_abi', ''),
                "Test Case": f"{item['test_class']}#{item['test_method']}",
                "Final Status": final_status,
                "Stability": stability,
                "Cluster (AI)": ai_cause,
                "Owner": owner,
                "Bug ID": bug_id,
                "Run History": hist_str,
                "Web Link": web_link,
                "Error Message": f_obj.error_message if f_obj else ""
            })
            
    # Sort: Persistent First
    cons_rows.sort(key=lambda x: 0 if x['Final Status'] == 'PERSISTENT' else 1)
    df_consolidated = pd.DataFrame(cons_rows)
    
    # --- Sheet 3: Run Details ---
    # We can reuse the runs from report_data
    runs_rows = []
    all_runs = []
    for s in report_data['suites']:
        all_runs.extend(s['runs'])
    all_runs.sort(key=lambda x: x.start_time)
    
    for r in all_runs:
         runs_rows.append({
            "Run ID": r.id,
            "Suite": r.test_suite_name,
            "Derived Suite": next((s['suite_name'] for s in report_data['suites'] if r in s['runs']), "Unknown"),
            "Device": r.device_fingerprint,
            "Start Time": clean_time(r.start_display), # Use clean time format
            "Total (Pass+Fail)": (r.passed_tests or 0) + (r.failed_tests or 0),
            "Passed": r.passed_tests,
            "Failed": r.failed_tests
        })
    df_runs = pd.DataFrame(runs_rows)

    # 3. Write Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Dashboard
        # Write Summary Rows
        start_row_kpi = 2 # Leave row 1 for Title
        df_summary_kpi.to_excel(writer, sheet_name='Dashboard', index=False, header=False, startrow=start_row_kpi)
        
        # Write Suite Table below
        start_row_suite = start_row_kpi + len(df_summary_kpi) + 2
        df_summary_suites.to_excel(writer, sheet_name='Dashboard', index=False, startrow=start_row_suite)
        
        # Sheet 2 & 3
        df_consolidated.to_excel(writer, sheet_name='Consolidated Analysis', index=False)
        df_runs.to_excel(writer, sheet_name='Run Details', index=False)
        
        # 4. Formatting (OpenPyXL)
        wb = writer.book
        
        # === Sheet 1 Styling ===
        ws_dash = wb['Dashboard']
        
        # 1. Title
        ws_dash['A1'] = "GMS Submission Report"
        ws_dash['A1'].font = Font(size=18, bold=True, color="2C3E50")
        ws_dash.merge_cells('A1:B1')
        
        # 2. KPI Section Styling
        # Columns widths
        ws_dash.column_dimensions['A'].width = 30
        ws_dash.column_dimensions['B'].width = 60
        
        # KPI Labels Bold
        for row in range(start_row_kpi + 1, start_row_kpi + len(df_summary_kpi) + 1):
            cell_key = ws_dash.cell(row=row, column=1)
            cell_val = ws_dash.cell(row=row, column=2)
            cell_key.font = Font(bold=True)
            cell_val.alignment = Alignment(horizontal='left')
            
            # Special Highlights
            if cell_key.value == "Overall Health":
                 cell_key.font = Font(size=14, bold=True, color="2980B9")
                 
            if cell_key.value == "Active Failures (Persistent)":
                cell_key.font = Font(color="C0392B", bold=True)
                cell_val.font = Font(color="C0392B", bold=True)
                
            if cell_key.value == "Pass Rate (Effective)":
                 cell_val.font = Font(size=12, bold=True, color="27AE60")

        # 3. Suite Table Styling
        # Headers (Row = start_row_suite + 1 because openpyxl is 1-based and to_excel wrote header there)
        header_row = start_row_suite + 1
        
        # Define Styles
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid") # Dark Blue
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Apply Header Style
        for col_idx, col_name in enumerate(suite_headers, 1):
            cell = ws_dash.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            # Set width for suite table columns
            ws_dash.column_dimensions[get_column_letter(col_idx)].width = 20

        # Apply Data Rows Style
        for r_idx in range(len(suite_rows)):
            row_num = header_row + 1 + r_idx
            for c_idx in range(len(suite_headers)):
                cell = ws_dash.cell(row=row_num, column=c_idx+1)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Pass Rate Color Scale
                if c_idx == 4: # Pass Rate Column
                    try:
                        val = float(cell.value.replace('%', ''))
                        if val == 100:
                            cell.font = Font(color="27AE60", bold=True) # Green
                        elif val < 90:
                            cell.font = Font(color="C0392B", bold=True) # Red
                    except: pass

        # === Sheet 2 Styling ===
        ws_cons = wb['Consolidated Analysis']
        ws_cons.column_dimensions['A'].width = 15 # Suite
        ws_cons.column_dimensions['B'].width = 30 # Module
        ws_cons.column_dimensions['C'].width = 15 # ABI
        ws_cons.column_dimensions['D'].width = 50 # Test Case
        ws_cons.column_dimensions['E'].width = 15 # Status
        ws_cons.column_dimensions['L'].width = 60 # Error Msg
        
        # Header Style for Sheet 2
        for cell in ws_cons[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Conditional Formatting for Status
        red_font = Font(color="9C0006")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_font = Font(color="006100")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        status_col_idx = 5 # E column (1-based index)
        link_col_idx = 11 # K column
        
        for row in ws_cons.iter_rows(min_row=2, max_row=ws_cons.max_row):
            status_cell = row[status_col_idx-1]
            if status_cell.value == "PERSISTENT":
                status_cell.fill = red_fill
                status_cell.font = red_font
            elif status_cell.value == "RECOVERED":
                status_cell.fill = green_fill
                status_cell.font = green_font
                
            # Hyperlink for Web Link
            link_cell = row[link_col_idx-1]
            if link_cell.value and link_cell.value.startswith("http"):
                link_cell.hyperlink = link_cell.value
                link_cell.value = "View Log"
                link_cell.font = Font(color="0000FF", underline="single")
        
        # Freeze panes
        ws_cons.freeze_panes = 'A2'
        
        # === Sheet 3 Styling ===
        ws_run = wb['Run Details']
        for cell in ws_run[1]:
            cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid") # Grey
            cell.font = Font(color="FFFFFF", bold=True)
        ws_run.column_dimensions['B'].width = 25 # Device
        ws_run.column_dimensions['E'].width = 20 # Time

    output.seek(0)
    
    filename = f"Submission_Report_v2_{sub.name}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
